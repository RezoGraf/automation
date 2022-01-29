from re import S
import time
import keyboard
from logging import error
import pyautogui
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl import load_workbook
import config

username = config.username
password = config.password


def  from_excel():
    wb = load_workbook("price2.xlsx")  # Work Book
    ws = wb.get_sheet_by_name('1')  # Work Sheet
    # column = ws['A']  # Column
    # column2 = ws['B']
    # print(column.value)
    s = []
    k = []
    list_data = []
    for i in ws['A']:
        s.append(i.value)
    for i in ws['B']:
        k.append(str(i.value))
    for i in range(len(s)):
        v = [s[i], k[i]]
        list_data.append(v)
    # print(s[1])
    # print(len(s))
    # print(type(s))
    # print(list_data)
    return list_data
    # for i in range(len(s)):
    #     b = b + data(s[i], k[i])
    # print(b)
    # print(column2.value)
    # for x in range(len(column)):
    #     column_list = [column[x].value,column2[x].value]
    # # column_list = [column[x].value, column2[x].value for x in range(len(column))]
    # print(column_list)


def hi(text, summa, interval=0.1):
    pyautogui.click(800, 1100)
    # pyautogui.click()
    time.sleep(1)
    keyboard.press('tab')
    keyboard.write(text, interval)
    keyboard.press('tab')
    time.sleep(1)
    pyautogui.press('delete')
    pyautogui.write(summa, interval)
    pyautogui.click(1900, 790)
    # pyautogui.press('\n')
    time.sleep(1)
    # pyautogui.click(800, 1140)
    print('finished')
    


def logon():
    from_excel()
    driver = webdriver.Chrome("chromedriver")
    driver.get("https://nok.rosminzdrav.ru/home/index")
    # найти поле имени пользователя / электронной почты и отправить само имя пользователя в поле ввода
    driver.find_element_by_id("login").send_keys(username)
    # найти поле ввода пароля и также вставить пароль
    driver.find_element_by_id("password").send_keys(password)
    # нажмите кнопку входа в систему
    # driver.find_element_by_name("submit").click()
    driver.find_element_by_xpath("//button[@type='submit']")
    # driver.find_element_by_css_selector("#login > input[type='text']").click()
    keyboard.press('enter')
    keyboard.press('F11')
    time.sleep(5)
    # driver.get("https://nok.rosminzdrav.ru/MO/index/14387")
    pyautogui.click(190, 505)
    time.sleep(5)
    # driver.find_element_by_class_name("k-button k-button-icontext k-grid-add']")

    # driver.find_element_by_class_name("k-button k-button-icontext k-grid-add")
    list_data = {}
    for i in list_data:
        hi(i[0], i[1])
    # hi('usluga2', '525')
    # ждем завершения состояния готовности
    WebDriverWait(driver=driver, timeout=10).until(
        lambda x: x.execute_script("return document.readyState === 'complete'")
    )
    error_message = "Неверный логин/пароль!"

    # получаем ошибки (если есть)
    errors = driver.find_elements_by_class_name("text-danger")

    # при необходимости распечатать ошибки
    # для e в ошибках:
    #     print(e.text)
    # если мы находим это сообщение об ошибке в составе error, значит вход не выполнен
    if any(error_message in e.text for e in errors):
        print("[!] Login failed")
    else:
        print("[+] Login successful")

    # закрываем драйвер
    # driver.close()





if __name__ == '__main__':
    # hi('привет')
    # logon()
    time.sleep(10)
    list_data = from_excel()
    for i in list_data:
        # print(i[1])
        hi(i[0], i[1])
